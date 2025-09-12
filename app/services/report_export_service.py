"""
Report Export Service Module
Handles exporting project reports to various formats (PDF, Excel).
"""
import logging
from datetime import datetime
from typing import Dict, Any, BinaryIO
from io import BytesIO
import json

from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from app.schemas.project import ProjectReportResponse

logger = logging.getLogger(__name__)


class ReportExportService:
    """Service for exporting project reports to different formats."""
    
    def __init__(self):
        self.styles = getSampleStyleSheet()
        self._setup_custom_styles()
    
    def _setup_custom_styles(self):
        """Setup custom styles for PDF generation."""
        # Title style
        self.styles.add(ParagraphStyle(
            name='CustomTitle',
            parent=self.styles['Heading1'],
            fontSize=18,
            spaceAfter=30,
            alignment=TA_CENTER,
            textColor=colors.darkblue
        ))
        
        # Section header style
        self.styles.add(ParagraphStyle(
            name='SectionHeader',
            parent=self.styles['Heading2'],
            fontSize=14,
            spaceAfter=12,
            spaceBefore=20,
            textColor=colors.darkgreen
        ))
        
        # Content style
        self.styles.add(ParagraphStyle(
            name='Content',
            parent=self.styles['Normal'],
            fontSize=10,
            spaceAfter=6,
            leftIndent=20
        ))
    
    async def export_to_pdf(self, report: ProjectReportResponse) -> BytesIO:
        """
        Export project report to PDF format.
        
        Args:
            report: ProjectReportResponse object
            
        Returns:
            BytesIO containing the PDF data
        """
        try:
            buffer = BytesIO()
            doc = SimpleDocTemplate(
                buffer,
                pagesize=A4,
                rightMargin=72,
                leftMargin=72,
                topMargin=72,
                bottomMargin=18
            )
            
            # Build the PDF content
            story = []
            
            # Title
            title = f"Project Report: {report.project_name}"
            story.append(Paragraph(title, self.styles['CustomTitle']))
            story.append(Spacer(1, 20))
            
            # Project summary table
            summary_data = [
                ['Project ID', report.project_id],
                ['Project Name', report.project_name],
                ['Status', report.project_status.value],
                ['Total Sections', str(report.total_sections)],
                ['Completed Sections', str(report.completed_sections)],
                ['Generated At', report.generated_at.strftime('%Y-%m-%d %H:%M:%S UTC')]
            ]
            
            summary_table = Table(summary_data, colWidths=[2*inch, 4*inch])
            summary_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
                ('BACKGROUND', (1, 0), (1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(summary_table)
            story.append(Spacer(1, 30))
            
            # Sections
            story.append(Paragraph("Sections Detail", self.styles['SectionHeader']))
            
            for section_key, section_data in report.sections.items():
                section_title = f"Section {section_data.section_number}"
                story.append(Paragraph(section_title, self.styles['Heading3']))
                
                status = "✓ Completed" if section_data.completed else "○ Not Completed"
                story.append(Paragraph(f"Status: {status}", self.styles['Content']))
                
                if section_data.final_output and section_data.completed:
                    # Format the final output
                    output_text = self._format_json_for_pdf(section_data.final_output)
                    story.append(Paragraph("Final Output:", self.styles['Content']))
                    story.append(Paragraph(output_text, self.styles['Content']))
                else:
                    story.append(Paragraph("No final output available", self.styles['Content']))
                
                story.append(Spacer(1, 15))
            
            # Build PDF
            doc.build(story)
            buffer.seek(0)
            
            logger.info(f"Successfully generated PDF report for project {report.project_id}")
            return buffer
            
        except Exception as e:
            logger.error(f"Error generating PDF report: {str(e)}")
            raise RuntimeError(f"Failed to generate PDF report: {str(e)}")
    
    async def export_to_excel(self, report: ProjectReportResponse) -> BytesIO:
        """
        Export project report to Excel format.
        
        Args:
            report: ProjectReportResponse object
            
        Returns:
            BytesIO containing the Excel data
        """
        try:
            buffer = BytesIO()
            workbook = Workbook()
            
            # Remove default sheet and create custom sheets
            workbook.remove(workbook.active)
            
            # Create summary sheet
            summary_sheet = workbook.create_sheet("Project Summary")
            self._create_summary_sheet(summary_sheet, report)
            
            # Create sections overview sheet
            sections_sheet = workbook.create_sheet("Sections Overview")
            self._create_sections_overview_sheet(sections_sheet, report)
            
            # Create individual sheets for completed sections
            for section_key, section_data in report.sections.items():
                if section_data.completed and section_data.final_output:
                    sheet_name = f"Section {section_data.section_number}"
                    section_sheet = workbook.create_sheet(sheet_name)
                    self._create_section_detail_sheet(section_sheet, section_data)
            
            # Save to buffer
            workbook.save(buffer)
            buffer.seek(0)
            
            logger.info(f"Successfully generated Excel report for project {report.project_id}")
            return buffer
            
        except Exception as e:
            logger.error(f"Error generating Excel report: {str(e)}")
            raise RuntimeError(f"Failed to generate Excel report: {str(e)}")
    
    def _format_json_for_pdf(self, data: Dict[str, Any]) -> str:
        """Format JSON data for PDF display."""
        try:
            if isinstance(data, dict):
                formatted_lines = []
                for key, value in data.items():
                    if isinstance(value, (dict, list)):
                        formatted_lines.append(f"<b>{key}:</b> {json.dumps(value, indent=2)}")
                    else:
                        formatted_lines.append(f"<b>{key}:</b> {str(value)}")
                return "<br/>".join(formatted_lines)
            else:
                return str(data)
        except Exception:
            return str(data)
    
    def _create_summary_sheet(self, sheet, report: ProjectReportResponse):
        """Create the project summary sheet in Excel."""
        # Headers
        header_font = Font(bold=True, size=12)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Title
        sheet['A1'] = f"Project Report: {report.project_name}"
        sheet['A1'].font = Font(bold=True, size=16)
        sheet.merge_cells('A1:B1')
        
        # Summary data
        summary_data = [
            ('Project ID', report.project_id),
            ('Project Name', report.project_name),
            ('Status', report.project_status.value),
            ('Total Sections', report.total_sections),
            ('Completed Sections', report.completed_sections),
            ('Generated At', report.generated_at.strftime('%Y-%m-%d %H:%M:%S UTC'))
        ]
        
        row = 3
        for label, value in summary_data:
            sheet[f'A{row}'] = label
            sheet[f'A{row}'].font = header_font
            sheet[f'B{row}'] = value
            row += 1
        
        # Auto-adjust column widths
        for column in sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[column_letter].width = adjusted_width
    
    def _create_sections_overview_sheet(self, sheet, report: ProjectReportResponse):
        """Create the sections overview sheet in Excel."""
        # Headers
        headers = ['Section Number', 'Status', 'Completed', 'Has Final Output']
        header_font = Font(bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Data rows
        row = 2
        for section_key, section_data in report.sections.items():
            sheet.cell(row=row, column=1, value=section_data.section_number)
            sheet.cell(row=row, column=2, value="Completed" if section_data.completed else "Not Completed")
            sheet.cell(row=row, column=3, value="Yes" if section_data.completed else "No")
            sheet.cell(row=row, column=4, value="Yes" if section_data.final_output else "No")
            row += 1
        
        # Auto-adjust column widths
        for column in sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[column_letter].width = adjusted_width
    
    def _create_section_detail_sheet(self, sheet, section_data):
        """Create a detailed sheet for an individual section."""
        # Title
        sheet['A1'] = f"Section {section_data.section_number} - Final Output"
        sheet['A1'].font = Font(bold=True, size=14)
        
        # Final output data
        if section_data.final_output:
            row = 3
            for key, value in section_data.final_output.items():
                sheet[f'A{row}'] = key
                sheet[f'A{row}'].font = Font(bold=True)
                
                if isinstance(value, (dict, list)):
                    sheet[f'B{row}'] = json.dumps(value, indent=2)
                else:
                    sheet[f'B{row}'] = str(value)
                
                row += 1
        
        # Auto-adjust column widths
        for column in sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 100)
            sheet.column_dimensions[column_letter].width = adjusted_width
    
    async def health_check(self) -> Dict[str, Any]:
        """
        Check the health of the report export service.
        
        Returns:
            Dict containing health status information
        """
        try:
            # Test basic functionality
            test_buffer = BytesIO()
            doc = SimpleDocTemplate(test_buffer, pagesize=letter)
            doc.build([Paragraph("Test", self.styles['Normal'])])
            
            return {
                "status": "healthy",
                "service": "report_export_service"
            }
        except Exception as e:
            return {
                "status": "unhealthy",
                "error": str(e),
                "service": "report_export_service"
            }
